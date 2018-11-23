import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from matplotlib import style, pyplot as plt

import numpy as np 


thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

wb = None

redFill = PatternFill(start_color='EE1111',
                    end_color='EE1111',
                    fill_type='solid')

def _reverse_dict_values(data):
    new_res = {}
    
    for key, val in data.items():

        for index in val:
            if index not in new_res:
                new_res.setdefault(index, [key])
                
            else:
                new_res[index].append(key)

    slices = [int(0)]*len(new_res)
    labels = ["None"]*len(new_res)
    explode = [int(0)]*len(new_res)
    
    for i, items in enumerate(new_res.items()):

        if "GPL" in items[0]:
            explode[i] = 0.15
        slices[i] = len(items[1])
        labels[i] = items[0]

    return (labels, slices, explode, new_res)




def _draw_chart(labels, slices, explode, new_res):

    fig, ax= plt.subplots(figsize=(4,4))
    plt.subplots_adjust(bottom=0.3)

    #plt.title("Licenses of Dependency Components")
    plt.gca().axis("equal")

    patches, texts = pie = plt.pie(slices, startangle=5)


    bbox_props = dict(boxstyle="square,pad=0.3", fc="w", ec="k", lw=0.72)
    arrowprops=dict(arrowstyle="-",connectionstyle="angle,angleA=0,angleB=90")
    kw = dict(xycoords='data',textcoords='data',arrowprops=dict(arrowstyle="-"), 
              bbox=bbox_props, zorder=0, va="center")

    for i, p in enumerate(patches):
        ang = (p.theta2 - p.theta1)/2. + p.theta1
        y = np.sin(np.deg2rad(ang))
        x = np.cos(np.deg2rad(ang))
        horizontalalignment = {-1: "right", 1: "left"}[int(np.sign(x))]
        connectionstyle = "angle,angleA=0,angleB={}".format(ang)
        kw["arrowprops"].update({"connectionstyle": connectionstyle})
        ax.annotate("\n".join(list(new_res.values())[i]), xy=(x, y), xytext=(1.35*np.sign(x), 1.4*y),
                     horizontalalignment=horizontalalignment, **kw)
        
    plt.legend(pie[0],labels, loc="center", bbox_to_anchor=(0.5,-0.2))
    plt.show()


'''
    Store final result in the excel format.

'''
def generateFinalReport(res_dict, result_type):

    if result_type == "graph":
        
        labels, slices, explode, new_res = _reverse_dict_values(res_dict)
        _draw_chart(labels, slices, explode, new_res)
    
    else:

        try:	
            wb = openpyxl.load_workbook('result.xlsx')
        except Exception:
            raise Exception("Please install openpyxl package!!\nCommand to install the package - (pip install openpyxl)")
        else:
            sheet = wb["Sheet1"]

            row = "C"
            col = 5
            
            for key, val in res_dict.items():
                sheet["%s"%(row+str(col))].border = thin_border
                sheet["%s"%(chr(ord(row)+1)+str(col))].border = thin_border
                sheet["%s"%(chr(ord(row)+2)+str(col))].border = thin_border
                sheet["%s"%(chr(ord(row)+3)+str(col))].border = thin_border
                
                sheet["%s"%(row+str(col))] = key
                sheet["%s"%(chr(ord(row)+1)+str(col))] = "%s"%(", ".join(v for v in val))

                if any(("%s"%v).find("General") != -1 for v in val):
                    sheet["%s"%(row+str(col))].fill = redFill
                    sheet["%s"%(chr(ord(row)+1)+str(col))].fill = redFill
                    sheet["%s"%(chr(ord(row)+2)+str(col))].fill = redFill
                    sheet["%s"%(chr(ord(row)+3)+str(col))].fill = redFill
                    
                col = col+1
                
            
            wb.save('result.xlsx')
        return True
    

if __name__ == "__main__":

    data = {
        "FOXopen": [
            "GPL License"
        ],
        "semver": [
            "MIT License"
        ],
        "logging-log4net": [
            "Apache License"
        ],
        "registry-url": [
            "Unlicensed"
        ]
    }
    generateFinalReport(data, "graph") 
